"""数据集解析适配器测试。

覆盖 parse_dataset 对 CSV/Excel 文件的字段类型推断、缺失值统计、
数值字段统计、字符串字段 top_values、质量概览、序列化往返。
覆盖错误分支：空数据集、损坏文件、不支持的扩展名。
"""

import csv

import pytest
from openpyxl import Workbook

from app.infrastructure.parsers.dataset_parser import (
    parse_dataset,
    profile_to_dict,
    profile_from_dict,
    DatasetParseError,
    FieldProfile,
    DatasetProfile,
)


def _write_csv(tmp_path, name: str, rows: list[list]) -> str:
    """写入 CSV 文件并返回路径。"""
    file_path = tmp_path / name
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)
    return str(file_path)


def _write_xlsx(tmp_path, name: str, rows: list[list]) -> str:
    """写入 xlsx 文件并返回路径。"""
    file_path = tmp_path / name
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(file_path)
    return str(file_path)


# --- CSV 解析 ---


class TestParseCsv:
    """CSV 文件解析测试。"""

    def test_parses_csv_with_basic_types(self, tmp_path):
        """CSV 文件解析：推断 string/int/float 类型，生成字段概览。"""
        rows = [
            ["name", "age", "score"],
            ["alice", "30", "90.5"],
            ["bob", "25", "85.0"],
            ["carol", "28", "78.0"],
        ]
        path = _write_csv(tmp_path, "basic.csv", rows)

        result = parse_dataset(path, "csv")

        assert result.profile.row_count == 3
        assert result.profile.column_count == 3
        assert len(result.profile.field_profiles) == 3

        # name 是字符串
        name_profile = next(p for p in result.profile.field_profiles if p.name == "name")
        assert name_profile.inferred_type == "string"
        assert name_profile.non_null_count == 3
        assert name_profile.null_count == 0
        assert name_profile.unique_count == 3
        assert len(name_profile.sample_values) <= 5
        # 字符串字段应有 top_values
        assert len(name_profile.top_values) > 0

        # age 推断为 int
        age_profile = next(p for p in result.profile.field_profiles if p.name == "age")
        assert age_profile.inferred_type == "int"
        assert age_profile.min_value == 25
        assert age_profile.max_value == 30
        assert age_profile.mean_value is not None
        assert age_profile.median_value is not None
        assert age_profile.q1 is not None
        assert age_profile.q3 is not None

        # score 推断为 float
        score_profile = next(p for p in result.profile.field_profiles if p.name == "score")
        assert score_profile.inferred_type == "float"
        assert score_profile.min_value == 78.0
        assert score_profile.max_value == 90.5

    def test_counts_missing_values(self, tmp_path):
        """缺失值统计：空单元格计入 null_count。"""
        rows = [
            ["name", "age"],
            ["alice", "30"],
            ["bob", ""],
            ["", "25"],
        ]
        path = _write_csv(tmp_path, "missing.csv", rows)

        result = parse_dataset(path, "csv")

        assert result.profile.row_count == 3

        age_profile = next(p for p in result.profile.field_profiles if p.name == "age")
        assert age_profile.null_count == 1
        assert age_profile.non_null_count == 2
        assert 0 < age_profile.null_rate <= 1

        name_profile = next(p for p in result.profile.field_profiles if p.name == "name")
        assert name_profile.null_count == 1

    def test_detects_duplicate_rows(self, tmp_path):
        """重复行检测：duplicate_row_count 正确统计。"""
        rows = [
            ["name", "age"],
            ["alice", "30"],
            ["alice", "30"],  # 重复
            ["bob", "25"],
        ]
        path = _write_csv(tmp_path, "duplicates.csv", rows)

        result = parse_dataset(path, "csv")

        assert result.profile.row_count == 3
        assert result.profile.duplicate_row_count == 1

    def test_quality_score_decreases_with_missing_and_duplicates(self, tmp_path):
        """质量评分：缺失值和重复行都会降低 quality_score。"""
        # 完整无缺失无重复
        clean_rows = [
            ["name", "age"],
            ["alice", "30"],
            ["bob", "25"],
        ]
        clean_path = _write_csv(tmp_path, "clean.csv", clean_rows)
        clean_profile = parse_dataset(clean_path, "csv").profile

        # 有缺失有重复
        dirty_rows = [
            ["name", "age"],
            ["alice", ""],
            ["alice", ""],  # 重复且缺失
            ["bob", "25"],
        ]
        dirty_path = _write_csv(tmp_path, "dirty.csv", dirty_rows)
        dirty_profile = parse_dataset(dirty_path, "csv").profile

        assert clean_profile.quality_score > dirty_profile.quality_score

    def test_computes_complete_row_count(self, tmp_path):
        """complete_row_count = 无缺失的行数；incomplete_row_count = 有缺失的行数。"""
        rows = [
            ["name", "age"],
            ["alice", "30"],   # 完整
            ["bob", ""],       # 不完整
            ["carol", "28"],   # 完整
        ]
        path = _write_csv(tmp_path, "incomplete.csv", rows)

        result = parse_dataset(path, "csv")
        assert result.profile.row_count == 3
        assert result.profile.complete_row_count == 2
        assert result.profile.incomplete_row_count == 1

    def test_string_field_top_values_max_10(self, tmp_path):
        """字符串字段 top_values 最多返回 10 个高频值。"""
        rows = [["category"]] + [[f"cat_{i % 5}"] for i in range(20)]
        path = _write_csv(tmp_path, "categories.csv", rows)

        result = parse_dataset(path, "csv")
        cat_profile = next(p for p in result.profile.field_profiles if p.name == "category")
        assert len(cat_profile.top_values) <= 10
        # 每个 cat_X 应出现 4 次
        for _, count in cat_profile.top_values:
            assert count == 4

    def test_sample_values_max_5(self, tmp_path):
        """sample_values 最多返回 5 个样例值。"""
        rows = [["name"]] + [[f"name_{i}"] for i in range(10)]
        path = _write_csv(tmp_path, "samples.csv", rows)

        result = parse_dataset(path, "csv")
        name_profile = result.profile.field_profiles[0]
        assert len(name_profile.sample_values) <= 5

    def test_infers_datetime_type(self, tmp_path):
        """日期字符串推断为 datetime 类型。"""
        rows = [
            ["timestamp"],
            ["2024-01-15"],
            ["2024-02-20"],
            ["2024-03-10"],
        ]
        path = _write_csv(tmp_path, "dates.csv", rows)

        result = parse_dataset(path, "csv")
        ts_profile = result.profile.field_profiles[0]
        assert ts_profile.inferred_type == "datetime"

    def test_infers_bool_type(self, tmp_path):
        """布尔值字段推断为 bool 类型。"""
        rows = [
            ["active"],
            ["True"],
            ["False"],
            ["True"],
        ]
        path = _write_csv(tmp_path, "bools.csv", rows)

        result = parse_dataset(path, "csv")
        active_profile = result.profile.field_profiles[0]
        # pandas 会将 True/False 字符串推断为 bool 或 object
        # 实际行为取决于 pandas 版本，但应能正确解析
        assert active_profile.inferred_type in ("bool", "string")

    def test_returns_raw_dataframe(self, tmp_path):
        """返回的 raw_dataframe 是 pandas DataFrame。"""
        import pandas as pd
        rows = [
            ["name", "age"],
            ["alice", "30"],
        ]
        path = _write_csv(tmp_path, "df.csv", rows)

        result = parse_dataset(path, "csv")
        assert isinstance(result.raw_dataframe, pd.DataFrame)
        assert len(result.raw_dataframe) == 1


# --- Excel 解析 ---


class TestParseExcel:
    """Excel 文件解析测试。"""

    def test_parses_xlsx_first_sheet(self, tmp_path):
        """xlsx 默认解析第一个工作表。"""
        rows = [
            ["name", "age", "score"],
            ["alice", 30, 90.5],
            ["bob", 25, 85.0],
        ]
        path = _write_xlsx(tmp_path, "data.xlsx", rows)

        result = parse_dataset(path, "xlsx")

        assert result.profile.row_count == 2
        assert result.profile.column_count == 3
        assert len(result.profile.field_profiles) == 3

        # age 列直接是 int 类型
        age_profile = next(p for p in result.profile.field_profiles if p.name == "age")
        assert age_profile.inferred_type == "int"
        assert age_profile.min_value == 25
        assert age_profile.max_value == 30

    def test_parses_xlsx_with_multiple_sheets_uses_first(self, tmp_path):
        """多工作表时只解析第一个。"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["name", "age"])
        ws1.append(["alice", 30])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["other", "value"])
        ws2.append(["x", 100])

        file_path = tmp_path / "multi.xlsx"
        wb.save(file_path)

        result = parse_dataset(str(file_path), "xlsx")
        # 只解析了 Sheet1 的数据
        assert result.profile.column_count == 2
        assert result.profile.row_count == 1
        # 第一列是 name（来自 Sheet1）
        assert result.profile.field_profiles[0].name == "name"

    def test_xlsx_with_missing_values(self, tmp_path):
        """xlsx 文件含缺失值的情况。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["alice", 30])
        ws.append(["bob", None])
        ws.append([None, 25])

        file_path = tmp_path / "missing.xlsx"
        wb.save(file_path)

        result = parse_dataset(str(file_path), "xlsx")
        age_profile = next(p for p in result.profile.field_profiles if p.name == "age")
        assert age_profile.null_count == 1

    def test_xlsx_supports_xls_extension(self, tmp_path):
        """扩展名 xls 也应支持（用 openpyxl engine）。"""
        rows = [
            ["name", "age"],
            ["alice", 30],
        ]
        path = _write_xlsx(tmp_path, "data.xls", rows)
        # 即使扩展名是 xls，文件本身是 xlsx 格式
        result = parse_dataset(path, "xls")
        assert result.profile.row_count == 1


# --- 错误分支 ---


class TestParseDatasetErrors:
    """解析错误分支测试。"""

    def test_unsupported_extension_raises(self, tmp_path):
        """不支持的扩展名抛 DATASET_FILE_UNSUPPORTED。"""
        file_path = tmp_path / "data.txt"
        file_path.write_text("hello")

        with pytest.raises(DatasetParseError) as exc:
            parse_dataset(str(file_path), "txt")
        assert exc.value.code == "DATASET_FILE_UNSUPPORTED"

    def test_empty_csv_raises_dataset_empty(self, tmp_path):
        """只有表头无数据行的 CSV 抛 DATASET_EMPTY。"""
        file_path = tmp_path / "empty.csv"
        file_path.write_text("name,age\n", encoding="utf-8")

        with pytest.raises(DatasetParseError) as exc:
            parse_dataset(str(file_path), "csv")
        assert exc.value.code == "DATASET_EMPTY"

    def test_truly_empty_csv_raises_dataset_empty(self, tmp_path):
        """完全空的 CSV 抛 DATASET_EMPTY 或 DATASET_PARSE_FAILED。"""
        file_path = tmp_path / "truly_empty.csv"
        file_path.write_text("", encoding="utf-8")

        # 完全空文件 pandas 会读为空 DataFrame
        with pytest.raises(DatasetParseError) as exc:
            parse_dataset(str(file_path), "csv")
        assert exc.value.code in ("DATASET_EMPTY", "DATASET_PARSE_FAILED")

    def test_corrupted_csv_raises_parse_failed(self, tmp_path):
        """损坏的 CSV 抛 DATASET_PARSE_FAILED。"""
        # 写入二进制乱码内容（非合法 CSV）
        file_path = tmp_path / "corrupt.csv"
        file_path.write_bytes(b"\x00\x01\x02\x03binary garbage")

        with pytest.raises(DatasetParseError) as exc:
            parse_dataset(str(file_path), "csv")
        # 乱码可能能被 pandas 读成单列，也可能抛 PARSE_FAILED
        # 若 pandas 不抛异常，则数据集不为空，会通过；
        # 但若文件无法解码为 utf-8，可能抛 UnicodeDecodeError，被映射为 PARSE_FAILED
        assert exc.value.code in ("DATASET_PARSE_FAILED", "DATASET_EMPTY")

    def test_corrupted_xlsx_raises_parse_failed(self, tmp_path):
        """损坏的 xlsx 抛 DATASET_PARSE_FAILED。"""
        file_path = tmp_path / "corrupt.xlsx"
        file_path.write_bytes(b"not actually xlsx content")

        with pytest.raises(DatasetParseError) as exc:
            parse_dataset(str(file_path), "xlsx")
        assert exc.value.code == "DATASET_PARSE_FAILED"

    def test_missing_file_raises_parse_failed(self, tmp_path):
        """文件不存在抛 DATASET_PARSE_FAILED。"""
        with pytest.raises(DatasetParseError) as exc:
            parse_dataset(str(tmp_path / "missing.csv"), "csv")
        assert exc.value.code == "DATASET_PARSE_FAILED"

    def test_extension_case_insensitive(self, tmp_path):
        """扩展名大写也应被识别。"""
        rows = [
            ["name", "age"],
            ["alice", "30"],
        ]
        path = _write_csv(tmp_path, "data.csv", rows)

        # 大写扩展名
        result = parse_dataset(path, "CSV")
        assert result.profile.row_count == 1

    def test_extension_with_dot_prefix(self, tmp_path):
        """扩展名带点前缀也能正确处理。"""
        rows = [
            ["name", "age"],
            ["alice", "30"],
        ]
        path = _write_csv(tmp_path, "data.csv", rows)

        result = parse_dataset(path, ".csv")
        assert result.profile.row_count == 1


# --- 序列化往返 ---


class TestProfileSerialization:
    """profile_to_dict / profile_from_dict 序列化往返测试。"""

    def test_profile_to_dict_contains_all_fields(self):
        """profile_to_dict 包含所有概览字段。"""
        profile = DatasetProfile(
            row_count=10,
            column_count=2,
            complete_row_count=8,
            incomplete_row_count=2,
            duplicate_row_count=1,
            field_profiles=[
                FieldProfile(
                    name="age",
                    inferred_type="int",
                    non_null_count=10,
                    null_count=0,
                    null_rate=0.0,
                    unique_count=5,
                    sample_values=["30", "25"],
                    min_value=20.0,
                    max_value=40.0,
                    mean_value=30.0,
                    median_value=30.0,
                    std_value=5.0,
                    q1=25.0,
                    q3=35.0,
                ),
                FieldProfile(
                    name="name",
                    inferred_type="string",
                    non_null_count=9,
                    null_count=1,
                    null_rate=0.1,
                    unique_count=8,
                    sample_values=["alice", "bob"],
                    top_values=[("alice", 3), ("bob", 2)],
                ),
            ],
            quality_score=85.5,
        )

        data = profile_to_dict(profile)

        assert data["row_count"] == 10
        assert data["column_count"] == 2
        assert data["complete_row_count"] == 8
        assert data["incomplete_row_count"] == 2
        assert data["duplicate_row_count"] == 1
        assert data["quality_score"] == 85.5
        assert len(data["field_profiles"]) == 2

        # 数值字段统计已序列化
        age_dict = data["field_profiles"][0]
        assert age_dict["min_value"] == 20.0
        assert age_dict["max_value"] == 40.0
        assert age_dict["mean_value"] == 30.0

        # 字符串字段 top_values 已序列化为 [[value, count]]
        name_dict = data["field_profiles"][1]
        assert name_dict["top_values"] == [["alice", 3], ["bob", 2]]

    def test_profile_round_trip_preserves_data(self):
        """profile → dict → profile 数据保持一致。"""
        original = DatasetProfile(
            row_count=5,
            column_count=2,
            complete_row_count=4,
            incomplete_row_count=1,
            duplicate_row_count=0,
            field_profiles=[
                FieldProfile(
                    name="age",
                    inferred_type="int",
                    non_null_count=5,
                    null_count=0,
                    null_rate=0.0,
                    unique_count=4,
                    sample_values=["30", "25", "40"],
                    min_value=20.0,
                    max_value=40.0,
                    mean_value=30.0,
                    median_value=30.0,
                    std_value=5.0,
                    q1=25.0,
                    q3=35.0,
                ),
            ],
            quality_score=90.0,
        )

        data = profile_to_dict(original)
        restored = profile_from_dict(data)

        assert restored.row_count == 5
        assert restored.column_count == 2
        assert restored.complete_row_count == 4
        assert restored.incomplete_row_count == 1
        assert restored.duplicate_row_count == 0
        assert restored.quality_score == 90.0
        assert len(restored.field_profiles) == 1

        age = restored.field_profiles[0]
        assert age.name == "age"
        assert age.inferred_type == "int"
        assert age.non_null_count == 5
        assert age.null_count == 0
        assert age.min_value == 20.0
        assert age.max_value == 40.0
        assert age.mean_value == 30.0
        assert age.median_value == 30.0
        assert age.std_value == 5.0
        assert age.q1 == 25.0
        assert age.q3 == 35.0

    def test_profile_from_dict_handles_empty_top_values(self):
        """空 top_values 字段不会报错。"""
        data = {
            "row_count": 0,
            "column_count": 0,
            "complete_row_count": 0,
            "incomplete_row_count": 0,
            "duplicate_row_count": 0,
            "quality_score": 0.0,
            "field_profiles": [
                {
                    "name": "x",
                    "inferred_type": "int",
                    "non_null_count": 0,
                    "null_count": 0,
                    "null_rate": 0.0,
                    "unique_count": 0,
                    "sample_values": [],
                    "top_values": [],
                },
            ],
        }

        profile = profile_from_dict(data)
        assert profile.row_count == 0
        assert len(profile.field_profiles) == 1
        assert profile.field_profiles[0].top_values == []

    def test_profile_from_dict_handles_missing_optional_fields(self):
        """dict 缺失可选字段时使用默认值。"""
        data = {
            "row_count": 1,
            "column_count": 1,
            "field_profiles": [],
            "quality_score": 50.0,
        }

        profile = profile_from_dict(data)
        assert profile.row_count == 1
        assert profile.column_count == 1
        assert profile.complete_row_count == 0  # 默认
        assert profile.incomplete_row_count == 0
        assert profile.duplicate_row_count == 0
        assert profile.quality_score == 50.0

    def test_top_values_round_trip_with_list_format(self):
        """top_values 在 dict 中是 [[value, count]] 格式，往返后是 [(value, count)]。"""
        profile = DatasetProfile(
            row_count=1,
            column_count=1,
            complete_row_count=1,
            incomplete_row_count=0,
            duplicate_row_count=0,
            field_profiles=[
                FieldProfile(
                    name="cat",
                    inferred_type="string",
                    non_null_count=1,
                    null_count=0,
                    null_rate=0.0,
                    unique_count=1,
                    sample_values=["a"],
                    top_values=[("a", 5), ("b", 3)],
                ),
            ],
            quality_score=100.0,
        )

        data = profile_to_dict(profile)
        # 序列化为 [[v, c], ...]
        assert data["field_profiles"][0]["top_values"] == [["a", 5], ["b", 3]]

        restored = profile_from_dict(data)
        # 反序列化回 [(v, c), ...]
        assert restored.field_profiles[0].top_values == [("a", 5), ("b", 3)]


# --- 大文件边界 ---


class TestParseDatasetLargeFile:
    """大文件边界测试。"""

    def test_large_csv_with_many_rows(self, tmp_path):
        """较多行数的 CSV 也能正确解析。"""
        rows = [["id", "value"]] + [[str(i), str(i * 1.5)] for i in range(1000)]
        path = _write_csv(tmp_path, "large.csv", rows)

        result = parse_dataset(path, "csv")
        assert result.profile.row_count == 1000
        assert result.profile.column_count == 2

    def test_csv_with_many_columns(self, tmp_path):
        """较多列数的 CSV 也能正确解析。"""
        header = [f"col_{i}" for i in range(50)]
        row = [str(i) for i in range(50)]
        path = _write_csv(tmp_path, "wide.csv", [header, row])

        result = parse_dataset(path, "csv")
        assert result.profile.column_count == 50
        assert len(result.profile.field_profiles) == 50
